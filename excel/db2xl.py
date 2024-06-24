#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# -*- encoding: utf-8 -*-
# Copyright (c) 2023 anqi.huang@outlook.com
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import json
import optparse
import os
from datetime import datetime

import dataset
import openpyxl
from openpyxl.styles import PatternFill

import tar_file


def dict2list(price_trend):
    list_of_keys = []
    list_of_values = []
    for key, val in dict(sorted(price_trend.items())).items():
        list_of_keys.append(key)
        list_of_values.append(val)

    return list_of_keys, list_of_values


def get_configs():
    configs = []

    configs_file = os.path.join(os.path.dirname(__file__), 'full5_only1.txt')

    if os.path.exists(configs_file):
        with open(configs_file, 'r', encoding="UTF-8") as f:
            for line in f.readlines():
                configs.append(line.strip())

    return configs


# 满五唯一
def full5_only1(deal_year, configs, title):
    if "满五年" == deal_year:
        for config in configs:
            if config.upper() in title.upper():
                return True

    return False


def save(districts, city, restrict, auto=0):
    # 转存到 excel
    if restrict:
        file_name = "{}-{}-lianjia".format(city, restrict)
    else:
        file_name = "{}-lianjia".format(city)

    # excel
    xl_file = os.path.join(os.path.dirname(__file__), file_name + '.xlsx')

    if os.path.exists(xl_file):
        os.remove(xl_file)

    workbook = openpyxl.Workbook()
    active = workbook.active
    workbook.remove(active)

    # db
    directory = os.path.join(os.path.dirname(__file__), "../data")
    if not os.path.exists(directory):
        os.makedirs(directory)

    if auto == 0:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        parent_dir = os.path.dirname(script_dir)
        tar_file.decompress_file(parent_dir, city, restrict)

    path = os.path.join(directory, file_name + '.db')
    database = dataset.connect('sqlite:///' + path)

    # configs
    configs = get_configs()

    total = 0
    up = 0
    down = 0
    now = 0

    for district in districts:
        table = database[district]
        sheet = workbook.create_sheet(district)
        sheet.append(('房屋ID',
                      '区域',
                      '商圈',
                      '小区',
                      '标题',
                      '总价',
                      '单价',
                      '涨(降)幅度',
                      '价格走势',
                      '房屋户型',
                      '所在楼层',
                      '建筑面积',
                      '年代',
                      '朝向',
                      '装修',
                      '梯户比例',
                      '配备电梯',
                      '挂牌时间',
                      '上次交易',
                      '房屋交易年限',
                      '交易权属',
                      '房屋用途',
                      '关注人数',
                      '带看人数',
                      '更新时间(第一次)',
                      '更新时间',
                      '房屋url'))
        col_range = sheet.max_column
        # 首行冻结
        sheet.freeze_panes = 'A2'
        # 隐藏某些列
        sheet.column_dimensions['A'].hidden = 1
        sheet.column_dimensions['Y'].hidden = 1
        sheet.column_dimensions['Z'].hidden = 1

        # 先保存到list里，是为了排序涨(降)价
        newerlist = list()

        for data in table.all():
            list_of_keys, list_of_values = dict2list(json.loads(data['price_trend'], object_hook=dict))
            trend = list_of_values[-1] - list_of_values[0]

            newerlist.append((data['house_id'],
                              data['district'],
                              data['bizcircle'],
                              data['xiaoqu'],
                              data['title'],
                              data['total_price'],
                              data['unit_price'],
                              trend,
                              data['price_trend'],
                              data['layout'],
                              data['flood'],
                              data['building_area'],
                              data['building_year'],
                              data['orientation'],
                              data['decoration'],
                              data['house_elevator'],
                              data['elevator'],
                              data['listing_time'],
                              data['last_deal'],
                              data['deal_year'],
                              data['house_characteristics'],
                              data['land_usage'],
                              data['follow_number'],
                              data['look_number'],
                              data['crawl_time'],
                              data['update_time'],
                              data['house_url'])
                             )

        # 按涨(降)价排序
        newerlist.sort(key=lambda a: a[7])

        for data in newerlist:
            trend = data[7]

            sheet.append(data)

            total += 1
            if trend > 0:
                up += 1
                # 涨价设置红色
                for col in range(1, col_range + 1):
                    cell = sheet.cell(sheet._current_row, col)
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
            elif trend < 0:
                down += 1
                # 降价设置绿色
                for col in range(1, col_range + 1):
                    cell = sheet.cell(sheet._current_row, col)
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")
            else:
                now += 1
                # 隐藏 持平 的房价
                sheet.row_dimensions[sheet._current_row].hidden = 1

            # if full5_only1(data['deal_year'], configs, data['title']):
            #     # 满五唯一
            #     for col in range(1, col_range + 1):
            #         cell = sheet.cell(sheet._current_row, col)
            #         cell.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type="solid")
            # else:
            #     sheet.row_dimensions[sheet._current_row].hidden = 1

    workbook.save(xl_file)
    database.close()

    if auto == 0:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        parent_dir = os.path.dirname(script_dir)
        tar_file.compress_file(parent_dir, city, restrict)

    msg = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    msg += ", 总房源 = {}(套), 涨价 = {}(套), 降价 = {}(套) , 持平 = {}(套)".format(total, up, down, now)
    return msg


def parseargs():
    usage = "usage: %prog [options] arg1 arg2"
    parser = optparse.OptionParser(usage=usage)

    option = optparse.OptionGroup(parser, "house scrapy crawl options")

    # 城市
    option.add_option("-c", "--city", dest="city", type="string",
                      help="city", default="sh-sf1a3a4a5p3-lianjia")
    option.add_option("-r", "--restrict", dest="restrict", type="string",
                      help="restrict", default="sf1a3a4a5p3")

    # 区域，如有多个则以/隔开
    option.add_option("-d", "--districts", dest="districts", type="string",
                      help="city districts", default="pudong/minhang/baoshan/songjiang/jiading/qingpu")

    parser.add_option_group(option)
    (options, args) = parser.parse_args()

    return (options, args)


if __name__ == '__main__':
    (options, args) = parseargs()
    city = options.city.strip()
    restrict = options.restrict.strip()
    district = options.districts.strip()
    districts = district.split("/")

    save(districts, city, restrict, 0)
